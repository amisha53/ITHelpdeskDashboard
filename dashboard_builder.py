import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ── Load data ──────────────────────────────────────────────────────────────────

def load_tickets(filepath):
    df = pd.read_csv(filepath, parse_dates=["date_opened", "date_closed"])
    df["resolution_hours"] = (
        (df["date_closed"] - df["date_opened"]).dt.total_seconds() / 3600
    )
    df["week"] = df["date_opened"].dt.isocalendar().week
    return df

fall = load_tickets("data/tickets_fall2023.csv")
spring = load_tickets("data/tickets_spring2024.csv")
fall["semester"] = "Fall 2023"
spring["semester"] = "Spring 2024"
combined = pd.concat([fall, spring], ignore_index=True)

# ── KPI calculations ───────────────────────────────────────────────────────────

def calc_kpis(df, label):
    total = len(df)
    avg_res = df["resolution_hours"].mean().round(2)
    fcr_rate = (df["first_contact_resolved"] == "Yes").mean() * 100
    top_category = df["category"].value_counts().idxmax()
    return {
        "Semester": label,
        "Total Tickets": total,
        "Avg Resolution Time (hrs)": avg_res,
        "First-Contact Resolution (%)": round(fcr_rate, 1),
        "Top Category": top_category,
    }

kpis = pd.DataFrame([calc_kpis(fall, "Fall 2023"), calc_kpis(spring, "Spring 2024")])

category_breakdown = (
    combined.groupby(["semester", "category"])
    .agg(count=("ticket_id", "count"), avg_hours=("resolution_hours", "mean"))
    .reset_index()
    .round(2)
)

weekly_volume = (
    combined.groupby(["semester", "week"])
    .size()
    .reset_index(name="ticket_count")
)

tech_workload = (
    combined.groupby(["semester", "technician"])
    .size()
    .reset_index(name="tickets_handled")
)

# ── Build Excel workbook ───────────────────────────────────────────────────────

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")
BODY_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(bold=True, name="Calibri", size=13)
BORDER      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def write_df(ws, df, start_row=1, title=None):
    if title:
        ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
        start_row += 1
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)
    style_header_row(ws, start_row, len(df.columns))
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value) + 4
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)
    return start_row + len(df) + 2

# Sheet 1 – KPI Summary
ws_kpi = wb.active
ws_kpi.title = "KPI Summary"
write_df(ws_kpi, kpis, start_row=1, title="Semester KPI Summary")

# Sheet 2 – Category Breakdown
ws_cat = wb.create_sheet("Category Breakdown")
write_df(ws_cat, category_breakdown, start_row=1, title="Tickets by Category")

# Sheet 3 – Weekly Volume
ws_vol = wb.create_sheet("Weekly Volume")
next_row = write_df(ws_vol, weekly_volume, start_row=1, title="Weekly Ticket Volume")

# Bar chart for weekly volume
chart = BarChart()
chart.title = "Weekly Ticket Volume"
chart.y_axis.title = "Tickets"
chart.x_axis.title = "Week"
data_ref = Reference(ws_vol, min_col=3, min_row=2, max_row=len(weekly_volume) + 2)
chart.add_data(data_ref, titles_from_data=False)
chart.series[0].title.v = "Ticket Count"
ws_vol.add_chart(chart, f"A{next_row}")

# Sheet 4 – Technician Workload
ws_tech = wb.create_sheet("Technician Workload")
write_df(ws_tech, tech_workload, start_row=1, title="Technician Workload")

output_path = "IT_Helpdesk_Dashboard.xlsx"
wb.save(output_path)
print(f"Dashboard saved to {output_path}")
